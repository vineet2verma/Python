from multiprocessing.connection import wait
from operator import contains
import smtplib
from email import encoders
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
import openpyxl
import warnings
import pandas as pd
import xlwings as xw
import msoffcrypto
import io
import datetime as dt
from pip import main
from pyparsing import col
from redash_dynamic_query import RedashDynamicQuery
from http.client import IM_USED
from time import sleep, time
import win32com.client as client
from PIL import ImageGrab
import datetime, os
from datetime import date
from datetime import datetime
from datetime import timedelta

file1 = 'C://Reporting//FW Meta Report//FW_Raw_Data//Redash_FW_mail_info.xlsx'
password = ' '
temp = io.BytesIO()
with open(file1, 'rb') as f:
    excel = msoffcrypto.OfficeFile(f)
    excel.load_key(password)
    excel.decrypt(temp)
df = pd.read_excel(temp,index_col=0,sheet_name='main_info')
df = df.dropna(how='all')
print(df.shape[0])

user_name = df[df['From'].str.contains('@',case=False).fillna(False)]['From']
paswd = df['Pswd'].dropna(how="all").tolist()
mainsend = df[df['Send To'].str.contains('@',case=False).fillna(False)]['Send To'].tolist()
strBcc = df[df['Send Bcc'].str.contains('@',case=False).fillna(False)]['Send Bcc'].tolist()

savelocation = df['save location'][1]
query_id = int(df['query id'][1])
col_formatting = df['col formatting'].tolist()
file_name = df['file name'][1]
chg_dir = df['chg dir/path'][1]
file_path = df['file_path'][1]
sh_name1 = df['sh_name1'][1]
copy_range1 = df['copy_range1'][1]
F_attach = df['File_Attch(Yes/No)'][1]

# Redash
redash = RedashDynamicQuery(
endpoint='https://redash.goibibo.com',
apikey='kdxvxHGOhFQcP5XMytV9Pn0kCzIEBk1lXS7pGpBW')
c_date = datetime.now().strftime("%Y_%m_%d_%I_%M_%p")
print("\n"+c_date)
os.chdir(savelocation)
query_id = query_id
result = redash.query(query_id)
gi= pd.DataFrame(result['query_result']['data']['rows'])
gi=gi[col_formatting]
file_name = file_name + c_date + ".csv"
gi.to_csv(file_name,index=False)

print("Redash Process Completed..\n\n")
os.chdir(chg_dir)
workbook_path = file_path

#  redash data 
neo_file = savelocation + file_name 
mapping = file_path

# df1 = pd.read_csv(neo_file,index_col=False)
# df2 = pd.read_excel(mapping,sheet_name='Mapping',index_col=False,engine='openpyxl')
# df2 = df2[['Hotelcode','Chain Name','Activation Status']]
# df3 = pd.merge(df1,df2,left_on='hotelcode',right_on='Hotelcode',how='inner').drop(['Hotelcode'],axis=1)

# print("\n Vlookup - Merging Process Done...\n ")

# wb = openpyxl.load_workbook(file_path)
# sh_raw = wb['Raw_File']
# maxrow = (sh_raw.max_row)
# maxcol = (sh_raw.max_column)

# print(f'Max Rows  :-  {maxrow}' )
# print(f'Max Columns :-  {maxcol}' )

# for r in sh_raw['A2:Q'+ str(maxrow)]:
#     for c in r:
#         c.value = None
# wb.save(file_path)

# print('\n Cleaning Done... \n')

# wwb = xw.Book(file_path)
# ssh = wwb.sheets['Raw_File']
# ssh['A2'].value = df3.values
# wwb.save()
# wwb.close()
# print('Data Paste Done')

# excel = client.Dispatch('Excel.Application')
# wb = excel.Workbooks.Open(workbook_path)
# print('\nRefreshing 2nd Excel Process Start..\n\n')
# sh2 = wb.Sheets['Raw_File']
# wb.RefreshAll()
# excel.CalculateUntilAsyncQueriesDone()
# wb.Save()
# sheet = wb.Sheets[sh_name1]

# copyrange= sheet.Range(copy_range1)
# copyrange.CopyPicture(Appearance=1, Format=2)
# ImageGrab.grabclipboard().save('paste.png')
# wb.Save()
# excel.Quit()
# img_path = os.getcwd() + '\\paste.png'

# print('Refresh Done & Save\n\n')
# date_str = dt.datetime.now().strftime("%d-%b-%y %I:%M:%S %p")
# strFrom = user_name[1]
# strTo = mainsend
# strBcc = strBcc
# msgRoot = MIMEMultipart('related')
# msgRoot['Subject'] = 'Neo Chain Room Night Trend '
# msgRoot['From'] = strFrom
# msgRoot['To'] =  ",".join(strTo)
# msgRoot['Cc'] = ",".join(strBcc)

# msgRoot.preamble = 'Multi-part message in MIME format.'
# msgAlternative = MIMEMultipart('alternative')
# msgRoot.attach(msgAlternative)
# msgText = MIMEText('Alternative plain text message.')
# msgAlternative.attach(msgText)

# msgText = MIMEText('''
# Hi,<br>
# Please find the attached <b style="color:red"> Top Neo Chain</b> production synopsis. <br>
# <div style="color:blue">This is an automated report</div>

# <img src="cid:image1"><br><br>

# <br><br>
# <p>Thanks & Regards,<br>
# Vineet Verma</p>
# ''', 'html')
# msgAlternative.attach(msgText)

# if F_attach == "Yes":
#     file_path = file_path
#     fileename = file_path.split("\\")[-1]
#     part = MIMEBase('application', "octet-stream")
#     part.set_payload(open(file_path, "rb").read())
#     encoders.encode_base64(part)
#     part.add_header('Content-Disposition', 'attachment; filename="%s"' %(fileename))
#     msgRoot.attach(part)

# fp = open('C:\\Reporting\\Neo Chains\\paste.png', 'rb')
# msgImage = MIMEImage(fp.read())
# fp.close()
# msgImage.add_header('Content-ID', '<image1>')
# msgRoot.attach(msgImage)
# smtp = smtplib.SMTP('smtp.outlook.office365.com', 587)
# smtp.starttls()
# smtp.login(user_name[1],paswd[0])
# smtp.sendmail(strFrom, (strTo+strBcc), msgRoot.as_string())
# del temp
# smtp.quit()



















# # SENDING MAIL BY SMTP SERVER-

# from dataclasses import replace
# from operator import index
# from statistics import mode
# import pandas as pd
# import xlwings as xw
# import numpy as np
# import warnings
# import smtplib, openpyxl
# from email import encoders
# from email.mime.multipart import MIMEMultipart
# from email.mime.text import MIMEText
# from email.mime.base import MIMEBase
# from email.mime.image import MIMEImage
# from matplotlib.pyplot import text
# import msoffcrypto
# import io
# import datetime as dt
# from redash_dynamic_query import RedashDynamicQuery
# from http.client import IM_USED
# from time import sleep, time
# from requests import head
# import win32com.client as client
# from PIL import ImageGrab
# import datetime, os
# from datetime import date
# from datetime import datetime
# from datetime import timedelta

# file1 = 'C:\\Reporting\\FW Meta Report\\Redash_Fw_Meta_info.xlsx' # information 
# df = pd.read_excel(file1,engine='openpyxl',index_col=0)
# savelocation = df.iloc[0][0]
# query_id = df.iloc[1][0]
# col_formatting = df.iloc[2].tolist()
# file_name = df.iloc[3][0]

# # Redash
# redash = RedashDynamicQuery(
# endpoint='https://redash.goibibo.com',
# apikey='kdxvxHGOhFQcP5XMytV9Pn0kCzIEBk1lXS7pGpBW')
# c_date = datetime.now().strftime("%Y_%m_%d_%I_%M_%p")
# print("\n"+c_date)
# os.chdir(savelocation)        
# result = redash.query(query_id)
# gi= pd.DataFrame(result['query_result']['data']['rows'])
# gi=gi[col_formatting]
# print(gi.shape[0])
# file_name = file_name + c_date + ".csv"
# gi.to_csv(file_name,index=False)


# # file4 = 'c:/Python/Redash working/NEO FILE/Neo_Rn_2022_07_14_03_20_PM.csv'
# # file3 = 'c:/Python/Redash working/test.xlsx'    # data paste file
# # wb = openpyxl.load_workbook(file3)
# # sh = wb['RN2']
# # lrow = sh.max_row
# # #  data clean
# # for row in sh['A2:O'+str(lrow)]:
# #     for cell in row:
# #         cell.value = ""

# # df4 = pd.read_csv(file4)
# # wb = xw.Book(file3)
# # sh = wb.sheets['RN2']
# # sh['A2'].value = df4.values

# # sh.range('P2:Q100').formula = sh.range('P2:Q2').formula
# # # range('RN2','P3:Q100').value = formula1

# # wb.save
# print("done")






