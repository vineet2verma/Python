# FUNCTION, BUTTON, TITLE, LABEL, ENTRY BOX, HEADER, COMBO BOX, CHECK BUTTON
# FRAME, MESSAGE BOX, LIST BOX
# TREE LIST UPDATE, READ EXCEL AS PER COL REQ.

from ast import Str
from ntpath import join
from time import sleep
import tkinter as tk
from tkinter import ANCHOR, BOTH, BOTTOM, CENTER, END, EXTENDED, LEFT, W, Checkbutton, Listbox, StringVar, ttk, messagebox
from tkinter import filedialog
from tkinter.filedialog import FileDialog
from tkinter.tix import COLUMN, Tree
from turtle import heading
from unicodedata import name
import openpyxl

from pandas import wide_to_long
from pyparsing import col
from setuptools import Command

win = tk.Tk()

# FRAME # NOT COMPLETE # NOT CLEAR
# frm_left = ttk.Frame(win,width=200,height=100)
# frm_left.pack(side=BOTTOM)

# fm1 = tk.Frame(win)
# fm1.pack(fill=tk.BOTH, expand=1)

# fm2 = tk.Frame(win)
# fm2.pack(fill=tk.BOTH, expand=1)

# WINDOW SIZE
win.maxsize(width= 800,height = 600)
win.minsize(width=800,height=400)

# CHECK BUTTON
chkbtn1 = tk.IntVar()
chkbtn2 = tk.IntVar()

select = ttk.Checkbutton(win,text="Male",variable=chkbtn1,onvalue=1,offvalue=0)
select.place(x=150,y=110)
select = ttk.Checkbutton(win,text="Female",variable=chkbtn2,onvalue=1,offvalue=0)
select.place(x=210,y=110)

# COMBO BOX
# v1 = StringVar()
# cbox1 = ttk.Combobox(win,width=27,textvariable=v1)
# cbox1['state'] = 'readonly'     # it's lock typing text in combo box
# cbox1['value'] = ['Jan','Feb','Mar']
# cbox1.current()    # set combo box 0 index value
# cbox1.place(x=150,y=80)

    
v1 = StringVar()
cbox1 = ttk.Combobox(win,width=27,textvariable=v1)
cbox1['state'] = 'readonly'     # it's lock typing text in combo box
cbox1['value'] = ['Jan','Feb','Mar']
cbox1.current()    # set combo box 0 index value
cbox1.place(x=150,y=80)

# LABEL
lbl1 = tk.Label(win, text="SMPLE USER FORM" ,font=40)    # TITLE 
lbl1.place(x=250,y=10)
lbl2 = tk.Label(win, text="User Name", bg="black", fg="white",font=6)
lbl2.place(x=30,y=40)
lbl4 = tk.Label(win, text="Combo Box",bg="black", fg="white",font=6)
lbl4.place(x=30,y=80)
lbl3 = tk.Label(win,text="List Box", bg="yellow", fg="black")
lbl3.place(x=580,y=40,width=100)
lbl5 = tk.Label(win,text="Check Box",bg="red", fg="white").place(x=30,y=110)


# MESSAGE BOX  #  WITH FUNCTION
def func2():
    # if ebox1_txt.get()  == "":
        # messagebox.showwarning("WARING","PLEASE FILL ENTRY BOX")
    # else:
        # messagebox.showinfo("Info Title", ebox1_txt.get())
        # messagebox.askyesno("WANT TO EXIT","EXIT FOR YES ELSE NO")
        win.destroy()
def fun3():                         # FOR DELETE IN LIST
    lst1.delete(ANCHOR)             # FOR DELETE SELECTED VALUES IN LIST
def fun4():                         # show selected value in list
    show.config(text=lst1.get(ANCHOR))

def fun5():                         # READ EXCEL FILE COLLEC DATA
    global name
    dic = {}
    name = filedialog.askopenfilename(title="File Selection Menu")
    wb = openpyxl.load_workbook(name)
    sheet = wb['ID']   
    for r,i in enumerate(range(sheet.max_row)):
        dic.update( {'ID' : r,'NO' : sheet.cell(row=i+1,column=1).value } )        
        tree.insert("",END,values = (dic['ID']+1,dic['NO']) )                 # ADDING VALUE IN TREE VIEW BY DICTIONARY
    

# TREE VIEW
mycolumn = ('ID', 'NO')
tree = ttk.Treeview(win,columns=mycolumn, show="headings")
# # ADDING COLUMN
tree.column('ID',width=50,anchor=CENTER)    # ANCHOR FOR TEXT ALIGMENT
tree.column('NO',width=50,anchor=CENTER)    # ANCHOR FOR TEXT ALIGMENT

# # ADDING HEADING
tree.heading('ID',text='ID_')       # HEADING COLUMN TEXT
tree.heading('NO',text='NO_')       # HEADING COLUMN TEXT

# # PACKING
# tree.pack(side=TOP ,fill=BOTH)
tree.place(x=580,y=70,height=180,width=150)



# # LIST BOX 
lst = []
# lst = ['Sunday','Monday','Tuesday','WEDNESDAY','THUSDAY','FRIDAY','SATURDAY']
# lst1 = Listbox(win,width=30)
# for r,i in enumerate(lst): lst1.insert(END,(f'{r+1} - {i}'))      # LAMBDA FOR LOOP FUNCTION
# lst1.place(x=580,y=70, width=100)

# ENTRY BOX
ebox1_txt = tk.StringVar()
ebox1 = tk.Entry(win, text="Enter Your Txt Here", bg="white", width=30,textvariable=ebox1_txt )
ebox1.place(x=150,y=40, height=25)

# BUTTON
btn1 = tk.Button(win, text="SUBMIT", bg="red")
btn1.place(x=150,y=220,height=25,width=110)
# btn_var = tk.StringVar()
btn2 = tk.Button(win,text="EXIT"  ,bg="red", command=func2)
btn2.place(x=300,y=220,width=110)
btn3 = tk.Button(win,text="DELETE", command=fun3,bg="YELLOW")
btn3.place(x=580,y=250)
btn4 = tk.Button(win,text="File Choose",command=fun5,bg="Yellow",fg="black")
btn4.place(x=300,y=250)

win.mainloop()












